"""
Convert the owner's master sheet into machine-readable rules.

Sources (C:\\Users\\NITTIN SHARMA\\OneDrive\\Desktop\\master sheet.xlsx):
  'Master Compliance'    -> 54 Companies Act controls (public + private)
  'SEBI PIT Compliance'  -> 42 PIT controls (listed / PIT-covered entities)

Outputs:
  rules/ca_master.json
  rules/pit_master.json

Same principle as build_lodr.py: only emit a computable due date when the
wording is unambiguous; grade everything else by confidence rather than
guessing, and give continuous duties no date at all.
"""
import openpyxl, json, io, re, sys
from collections import Counter

XLSX = r'C:\Users\NITTIN SHARMA\OneDrive\Desktop\master sheet.xlsx'

FY_END = {'month': 3, 'day': 31}
QUARTER_ENDS = [
    {'q': 'Q1', 'covers': 'Apr-Jun', 'month': 6,  'day': 30},
    {'q': 'Q2', 'covers': 'Jul-Sep', 'month': 9,  'day': 30},
    {'q': 'Q3', 'covers': 'Oct-Dec', 'month': 12, 'day': 31},
    {'q': 'Q4', 'covers': 'Jan-Mar', 'month': 3,  'day': 31},
]
HALF_YEAR_ENDS = [
    {'h': 'H1', 'covers': 'Apr-Sep', 'month': 9, 'day': 30},
    {'h': 'H2', 'covers': 'Oct-Mar', 'month': 3, 'day': 31},
]


def clean(v):
    if v is None:
        return None
    s = (str(v).replace('\u2013', '-').replace('\u2014', '-')
         .replace('\u2019', "'").replace('\u25b8', '>').strip())
    return s or None


def parse_timeline(text, freq):
    """Return (due_dict, confidence)."""
    t = (text or '').strip()
    tl = t.lower()
    f = (freq or '').strip().lower()

    if not t:
        return {'type': 'review'}, 'unknown'

    # Continuous duties get no date. Dating them manufactures false overdues.
    if re.match(r'^(continuous|ongoing|always|perpetual)', tl) or f.startswith('continuous'):
        return {'type': 'continuous'}, 'none'

    # "within N days of the AGM"  (AOC-4 is 30, MGT-7 is 60)
    m = re.search(r'within\s+(\d+)\s+days?\s+(of|from)\s+(the\s+)?agm', tl)
    if m:
        return {'type': 'agm_offset', 'days': int(m.group(1))}, 'exact'

    # "within N months of FY end"  (AGM is 6 months)
    m = re.search(r'within\s+(\d+)\s+months?\s+(of|from)\s+(the\s+)?(fy|financial year)\s*end', tl)
    if m:
        return {'type': 'fy_end_month_offset', 'months': int(m.group(1)),
                'anchorDate': FY_END}, 'exact'

    # "within N days of <event>"
    m = re.search(r'within\s+(\d+)\s+(working\s+)?days?\s+(of|from)\s+', tl)
    if m:
        return {'type': 'event_offset', 'days': int(m.group(1)),
                'workingDays': bool(m.group(2)), 'trigger': 'event'}, 'exact'

    # Fixed calendar dates: "by 30 September", "on or before 30th June"
    m = re.search(r'(?:by|before|on or before)\s+(\d{1,2})(?:st|nd|rd|th)?\s+'
                  r'(january|february|march|april|may|june|july|august|september|october|november|december)', tl)
    if m:
        months = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
                  'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
        return {'type': 'fixed_date', 'month': months[m.group(2)], 'day': int(m.group(1))}, 'exact'

    # Board-meeting cadence
    if 'minimum 4 meeting' in tl or 'min. 4' in tl or '120 days' in tl:
        return {'type': 'cadence', 'minPerYear': 4, 'maxGapDays': 120}, 'exact'

    if 'first board meeting' in tl:
        return {'type': 'annual', 'anchor': 'first_board_meeting', 'anchorDate': FY_END}, 'derived'

    # Cadence words with no offset — real, anchored, but the exact date is practice.
    if f.startswith('annual'):
        return {'type': 'annual', 'anchorDate': FY_END}, 'derived'
    if f.startswith('half'):
        return {'type': 'half_yearly', 'anchorDates': HALF_YEAR_ENDS}, 'derived'
    if f.startswith('quarter'):
        return {'type': 'quarterly', 'anchorDates': QUARTER_ENDS}, 'derived'
    if f.startswith('event') or f.startswith('recurring') or 'trigger' in f:
        return {'type': 'event', 'trigger': 'event'}, 'derived'

    # Immediate / point-of-event controls. These are standing obligations tied to
    # a trigger, not calendar items — giving them a due date would be wrong.
    if re.search(r'^(immediate|prompt|before|at event|at execution|at trade|on receipt|'
                 r'upon|at inception|real[- ]time|as it arises)', tl):
        return {'type': 'at_trigger', 'when': t}, 'trigger'
    if re.search(r'(before sharing|before receipt|before trading|before submission|'
                 r'at event inception|at execution|prompt update)', tl):
        return {'type': 'at_trigger', 'when': t}, 'trigger'
    if 'within prescribed timeline' in tl or 'as prescribed' in tl:
        return {'type': 'review'}, 'unknown'

    return {'type': 'review'}, 'unknown'


def parse_ca_applicability(text):
    """Companies Act sheet: 'Public + Private', 'Unlisted Public only', 'Private only'."""
    tl = (text or '').lower()
    if 'private only' in tl:
        return {'entityType': ['private', 'opc']}
    if 'unlisted public only' in tl:
        return {'entityType': ['public']}
    if 'public + private' in tl:
        return {'entityType': ['private', 'public', 'listed', 'opc', 'sec8']}
    return {'entityType': ['private', 'public', 'listed', 'opc', 'sec8']}


def build_ca(wb):
    ws = wb['Master Compliance']
    rows = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]
    out = []
    for r in rows:
        rid, appl, cat, title, prov, freq, timeline, detail, owner, evidence, risk = (
            [clean(x) for x in r[:11]] + [None] * (11 - len(r[:11])))
        due, conf = parse_timeline(timeline, freq)
        out.append({
            'id': 'CA-' + re.sub(r'[^A-Za-z0-9]+', '-', (prov or rid or '')).strip('-').upper()[:40] + '-' + str(rid),
            'sourceRow': rid,
            'law': 'Companies Act 2013',
            'regulation': prov,
            'category': cat,
            'title': title,
            'frequency': freq,
            'timelineText': timeline,
            'due': due,
            'dueConfidence': conf,
            'appliesTo': parse_ca_applicability(appl),
            'appliesToText': appl,
            'trigger': detail,
            'owner': owner,
            'evidence': evidence,
            'risk': (risk or 'medium'),
            'needsReview': conf == 'unknown',
        })
    return out


def build_pit(wb):
    ws = wb['SEBI PIT Compliance']
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]
    out = []
    for i, r in enumerate(rows, 1):
        reg, area, control, detail, appl, trigger, timeline, owner, evidence, risk, prio = (
            [clean(x) for x in r[:11]] + [None] * (11 - len(r[:11])))
        due, conf = parse_timeline(timeline, trigger)
        out.append({
            'id': 'PIT-' + re.sub(r'[^A-Za-z0-9]+', '-', (reg or '')).strip('-').upper()[:30] + '-' + str(i),
            'sourceRow': i,
            'law': 'SEBI PIT Regulations 2015',
            'regulation': reg,
            'area': area,
            'title': control,
            'detail': detail,
            'frequency': trigger,
            'timelineText': timeline,
            'due': due,
            'dueConfidence': conf,
            # PIT bites on listed / PIT-covered entities.
            'appliesTo': {'entityType': ['listed']},
            'appliesToText': appl,
            'owner': owner,
            'evidence': evidence,
            'risk': (risk or 'high'),
            'priority': prio,
            'needsReview': conf == 'unknown',
        })
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ca = build_ca(wb)
    pit = build_pit(wb)

    meta = {'source': 'master sheet.xlsx', 'fyEnd': FY_END,
            'quarterEnds': QUARTER_ENDS, 'halfYearEnds': HALF_YEAR_ENDS}

    io.open('rules/ca_master.json', 'w', encoding='utf-8').write(
        json.dumps({'meta': dict(meta, sheet='Master Compliance', count=len(ca)), 'rules': ca},
                   ensure_ascii=False, indent=1))
    io.open('rules/pit_master.json', 'w', encoding='utf-8').write(
        json.dumps({'meta': dict(meta, sheet='SEBI PIT Compliance', count=len(pit)), 'rules': pit},
                   ensure_ascii=False, indent=1))

    print(f'Companies Act : {len(ca)} controls')
    print('  confidence  :', dict(Counter(c['dueConfidence'] for c in ca)))
    print('  due types   :', dict(Counter(c['due']['type'] for c in ca)))
    print()
    print(f'SEBI PIT      : {len(pit)} controls')
    print('  confidence  :', dict(Counter(c['dueConfidence'] for c in pit)))
    print('  due types   :', dict(Counter(c['due']['type'] for c in pit)))


if __name__ == '__main__':
    main()
